#!/usr/bin/env python3
"""
shift_sleep_guard.py
Patrícia Menezes — Visaya Health Group
Analyzes nurse shift CSVs for insufficient rest gaps (< 11 h) and scores
weekly sleep risk per nurse. Outputs a color-coded Excel report.
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter


# ── Tuning ─────────────────────────────────────────────────────────────────────
MIN_REST_HOURS = 11   # Brazilian nursing regulation minimum
CRITICAL_HOURS = 6    # anything under this = HIGH risk
WARN_HOURS     = 8    # anything under this = MODERATE risk

COL_STATUS    = "Rest Status"
COL_VIOLATION = "Violation"


# ── Helpers ───────────────────────────────────────────────────────────────────

def load_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    required = {"nurse_name", "shift_date", "shift_start", "shift_end"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"CSV missing required columns: {sorted(missing)}")
    return df


def parse_dt(date_val, start_val, end_val):
    """
    Convert (date_str, start_str, end_str) naive times to UTC-aware datetimes.
    Naive times are treated as São Paulo local time (UTC-3).
    Returns (start_utc, end_utc) as tz-naive UTC floats (days since epoch).
    """
    date  = pd.to_datetime(date_val).date()
    start = pd.to_datetime(start_val, format="mixed").time()
    end   = pd.to_datetime(end_val,   format="mixed").time()

    def to_utc_naive(dt):
        loc = pd.Timestamp(dt).tz_localize("America/Sao_Paulo")
        return loc.tz_convert("UTC").tz_localize(None)

    start_utc = to_utc_naive(pd.Timestamp.combine(date, start))
    if end <= start:
        end_dt = pd.Timestamp.combine(date + pd.Timedelta(days=1), end)
    else:
        end_dt = pd.Timestamp.combine(date, end)
    return start_utc, to_utc_naive(end_dt)


def build_violations(df: pd.DataFrame):
    """
    For every consecutive pair of shifts per nurse, compute the gap in hours.
    Returns a DataFrame with one row per gap (not per shift).
    """
    rows = []
    for _, r in df.iterrows():
        s, e = parse_dt(r["shift_date"], r["shift_start"], r["shift_end"])
        rows.append({"nurse_name": r["nurse_name"],
                     "shift_date":  r["shift_date"],
                     "shift_start": r["shift_start"],
                     "shift_end":   r["shift_end"],
                     "start_dt": s,
                     "end_dt":   e})

    work = pd.DataFrame(rows).sort_values(["nurse_name", "start_dt"]).reset_index(drop=True)

    records = []
    for nurse, grp in work.groupby("nurse_name", sort=False):
        grp = grp.sort_values("start_dt").reset_index(drop=True)
        for i in range(1, len(grp)):
            gap_h = (grp.loc[i, "start_dt"] - grp.loc[i - 1, "end_dt"]).total_seconds() / 3600
            if gap_h < 0:
                continue          # malformed / overlapping, skip
            violating = gap_h < MIN_REST_HOURS
            records.append({
                "nurse_name":       nurse,
                "shift_date":       grp.loc[i, "shift_date"],
                "shift_start":     grp.loc[i, "shift_start"],
                "shift_end":       grp.loc[i, "shift_end"],
                "prev_shift_end":   grp.loc[i - 1, "end_dt"].strftime("%Y-%m-%d %H:%M"),
                "gap_hours":       round(gap_h, 2),
                "shortfall_hours": round(MIN_REST_HOURS - gap_h, 2),
                COL_VIOLATION:    "⚠️ YES" if violating else "✓ OK",
                COL_STATUS:        _severity(gap_h),
                "Break Hours":     round(gap_h, 2),
            })

    return pd.DataFrame(records)


def _severity(gap: float) -> str:
    if gap < CRITICAL_HOURS: return "🔴 HIGH"
    if gap < WARN_HOURS:     return "🟡 MODERATE"
    return "🟢 OK"


def weekly_scores(violations_df: pd.DataFrame) -> pd.DataFrame:
    """
    Per (nurse, ISO week) sleep risk score.
    Score = violations × 10 + avg_shortfall_h × 5
    """
    if violations_df.empty:
        return pd.DataFrame(columns=[
            "nurse_name", "week_start", "total_shifts",
            "violations", "avg_shortfall_h", "sleep_risk_score", "risk_level"
        ])

    v = violations_df[violations_df[COL_VIOLATION] == "⚠️ YES"].copy()
    v["week_start"] = (
        pd.to_datetime(v["shift_date"])
          .apply(lambda d: d - pd.Timedelta(days=d.weekday()))
    )

    weekly = (
        v.groupby(["nurse_name", "week_start"])
          .agg(violations=("nurse_name", "count"),
               avg_shortfall_h=("shortfall_hours", "mean"))
          .reset_index()
    )

    # Total shifts per (nurse, week) from the original violations frame
    all_shifts = (
        violations_df[["nurse_name", "shift_date"]].drop_duplicates()
        .assign(week_start=lambda d: (
            pd.to_datetime(d["shift_date"])
            .apply(lambda x: x - pd.Timedelta(days=x.weekday()))
        ))
        .groupby(["nurse_name", "week_start"])
        .size()
        .rename("total_shifts")
    )
    weekly = weekly.set_index(["nurse_name", "week_start"])
    weekly["total_shifts"] = all_shifts.reindex(weekly.index, fill_value=0)
    weekly = weekly.reset_index()

    weekly["sleep_risk_score"] = (
        weekly["violations"] * 10 +
        weekly["avg_shortfall_h"].fillna(0) * 5
    ).round(1)

    weekly["risk_level"] = weekly["sleep_risk_score"].apply(lambda s:
        "🔴 HIGH" if s >= 20 else ("🟡 MODERATE" if s >= 10 else "🟢 LOW")
    )

    return weekly.sort_values(["week_start", "sleep_risk_score"], ascending=[True, False])


# ── Excel writer ───────────────────────────────────────────────────────────────

def write_excel(violations_df, scores_df, out_path: Path):
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        violations_df.to_excel(w, sheet_name="Violations",        index=False)
        scores_df.to_excel(   w, sheet_name="Weekly Risk Score", index=False)

    wb = load_workbook(out_path)

    for sheet_name, df, status_col in [
        ("Violations",         violations_df, COL_STATUS),
        ("Weekly Risk Score",  scores_df,     "risk_level"),
    ]:
        _style_sheet(wb[sheet_name], df, status_col)

    wb.save(out_path)
    print(f"✅  Report → {out_path}")


def _style_sheet(ws, df, status_col):
    HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
    THIN_BORDER   = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin")
    )

    # Header row
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border   = THIN_BORDER

    # Data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border   = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
        # Bold nurse name column
        row[0].font = Font(bold=True)

    # Auto-width columns
    for col in ws.columns:
        longest = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(longest + 4, 50)

    # Conditional formatting on status column
    status_idx = None
    for i, h in enumerate(ws[1], 1):
        if h.value == status_col:
            status_idx = i
            break

    if status_idx:
        col_letter = get_column_letter(status_idx)
        last_row  = len(df) + 1

        cf_rules = [
            ("🔴 HIGH",    "FF0000"),
            ("🟡 MODERATE","FFBF00"),
            ("🟢 OK",     "00B050"),
            ("🟢 LOW",    "00B050"),
        ]
        for label, fg_hex in cf_rules:
            fill = PatternFill("solid", fgColor=fg_hex)
            dxf  = DifferentialStyle(fill=fill)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{last_row}",
                Rule(type="containsText", operator="containsText",
                     text=label, dxf=dxf)
            )


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Shift Sleep Guard — rest gap analyzer")
    ap.add_argument("--csv",   required=True, help="Path to nurse shift CSV")
    ap.add_argument("--out",  default="shift_sleep_report.xlsx",
                    help="Output Excel path")
    args = ap.parse_args()

    csv_path = Path(args.csv)
    out_path = Path(args.out)

    if not csv_path.exists():
        print(f"❌  CSV not found: {csv_path}")
        return 1

    print(f"📄  Loading {csv_path}")
    df = load_csv(csv_path)
    print(f"   {len(df)} rows, {df['nurse_name'].nunique()} nurses")

    print("🔍  Analyzing rest gaps…")
    violations = build_violations(df)

    if violations.empty:
        print("   No gaps found — all clear!")
    else:
        hi = (violations[COL_VIOLATION] == "⚠️ YES").sum()
        print(f"   {len(violations)} gaps analyzed, {hi} violations")

    print("📊  Computing weekly risk scores…")
    scores = weekly_scores(violations)

    print(f"💾  Writing {out_path}")
    write_excel(violations, scores, out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())